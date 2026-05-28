from __future__ import annotations

import json
import re
from copy import copy
from collections import Counter
from urllib.parse import parse_qsl, quote, unquote, urlencode, urlsplit, urlunsplit

from PyQt6.QtWidgets import (
    QFileDialog,
    QComboBox,
    QFormLayout,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from ui.common import bottom_action_bar

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None


class PlainTextEdit(QTextEdit):
    def insertFromMimeData(self, source) -> None:
        self.insertPlainText(source.text())


class UrlCodecTab(QWidget):
    def __init__(self, main) -> None:
        super().__init__()
        self.main = main
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        content = QWidget()
        layout = QVBoxLayout(content)
        self.input = PlainTextEdit()
        self.output = PlainTextEdit()
        encode = QPushButton("URL 인코딩")
        decode = QPushButton("URL 디코딩")
        copy = QPushButton("결과 복사")
        encode.clicked.connect(lambda: self.convert(True))
        decode.clicked.connect(lambda: self.convert(False))
        copy.clicked.connect(lambda: self.main.app.clipboard().setText(self.output.toPlainText()))
        layout.addWidget(QLabel("입력"))
        layout.addWidget(self.input, 1)
        layout.addWidget(QLabel("결과"))
        layout.addWidget(self.output, 1)
        root.addWidget(content, 1)
        root.addLayout(bottom_action_bar(encode, decode, copy))

    def convert(self, encode: bool) -> None:
        text = self.input.toPlainText()
        self.output.setPlainText(quote(text, safe="") if encode else unquote(text))


class UtmTab(QWidget):
    UTM_KEYS = ["utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content"]

    def __init__(self, main) -> None:
        super().__init__()
        self.main = main
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        content = QWidget()
        layout = QVBoxLayout(content)
        form = QFormLayout()
        self.url = QLineEdit()
        self.fields = {key: QLineEdit() for key in self.UTM_KEYS}
        form.addRow("URL", self.url)
        for key, field in self.fields.items():
            form.addRow(key, field)
        layout.addLayout(form)
        split = QPushButton("분해")
        build = QPushButton("조합")
        copy = QPushButton("URL 복사")
        split.clicked.connect(self.split_url)
        build.clicked.connect(self.build_url)
        copy.clicked.connect(lambda: self.main.app.clipboard().setText(self.url.text()))
        layout.addStretch(1)
        root.addWidget(content, 1)
        root.addLayout(bottom_action_bar(split, build, copy))

    def split_url(self) -> None:
        parts = urlsplit(self.url.text().strip())
        params = dict(parse_qsl(parts.query, keep_blank_values=True))
        for key, field in self.fields.items():
            field.setText(params.get(key, ""))

    def build_url(self) -> None:
        raw = self.url.text().strip()
        parts = urlsplit(raw)
        params = dict(parse_qsl(parts.query, keep_blank_values=True))
        for key, field in self.fields.items():
            value = field.text().strip()
            if value:
                params[key] = value
            else:
                params.pop(key, None)
        query = urlencode(params)
        self.url.setText(urlunsplit((parts.scheme, parts.netloc, parts.path, query, parts.fragment)))


class LineBreakTab(QWidget):
    EXAMPLES = {
        "줄바꿈 > 쉼표": ("hello\nworld", "hello, world"),
        "공백 > 줄바꿈": ("hello world foo", "hello\nworld\nfoo"),
        "줄바꿈 > 따옴표 목록": ("hello\nworld", "('hello', 'world')"),
        "공백 > 따옴표 목록": ("hello world foo", "('hello', 'world', 'foo')"),
    }

    def __init__(self, main) -> None:
        super().__init__()
        self.main = main
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        content = QWidget()
        layout = QVBoxLayout(content)
        self.mode = QComboBox()
        self.mode.addItems(list(self.EXAMPLES))
        self.mode.currentTextChanged.connect(self.update_example)
        self.before_example = QLabel()
        self.after_example = QLabel()
        for label in [self.before_example, self.after_example]:
            label.setObjectName("mutedText")
            label.setWordWrap(False)
            label.setFixedHeight(32)
            label.setStyleSheet("QLabel#mutedText { padding: 2px 6px; border: 1px solid rgba(107,114,128,0.28); border-radius: 6px; }")
        self.input = PlainTextEdit()
        self.output = PlainTextEdit()
        convert = QPushButton("변환")
        copy = QPushButton("결과 복사")
        convert.clicked.connect(self.convert)
        copy.clicked.connect(lambda: self.main.app.clipboard().setText(self.output.toPlainText()))
        row = QHBoxLayout()
        row.setSpacing(6)
        row.addWidget(self.mode)
        row.addWidget(self.before_example, 1)
        row.addWidget(self.after_example, 1)
        layout.addLayout(row)
        layout.addWidget(QLabel("입력"))
        layout.addWidget(self.input, 1)
        layout.addWidget(QLabel("결과"))
        layout.addWidget(self.output, 1)
        root.addWidget(content, 1)
        root.addLayout(bottom_action_bar(convert, copy))
        self.update_example(self.mode.currentText())

    def update_example(self, mode: str) -> None:
        before, after = self.EXAMPLES.get(mode, ("", ""))
        self.before_example.setText(before)
        self.after_example.setText(after)

    def convert(self) -> None:
        mode = self.mode.currentText()
        text = self.input.toPlainText().strip()
        items = [line.strip() for line in text.splitlines() if line.strip()] if mode.startswith("줄바꿈") else [part.strip() for part in text.split() if part.strip()]
        if "쉼표" in mode:
            result = ", ".join(items)
        elif mode == "공백 > 줄바꿈":
            result = "\n".join(items)
        else:
            result = "(" + ", ".join(f"'{item}'" for item in items) + ")"
        self.output.setPlainText(result)


class CaseCycleTab(QWidget):
    def __init__(self, main) -> None:
        super().__init__()
        self.main = main
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        content = QWidget()
        layout = QVBoxLayout(content)
        hint = QLabel("UPPER_CASE -> camelCase -> PascalCase 형식으로 순환 변환합니다.")
        hint.setObjectName("mutedText")
        layout.addWidget(hint)
        self.input = PlainTextEdit()
        self.output = PlainTextEdit()
        convert = QPushButton("순환 변환")
        copy = QPushButton("결과 복사")
        convert.clicked.connect(self.convert)
        copy.clicked.connect(lambda: self.main.app.clipboard().setText(self.output.toPlainText()))
        layout.addWidget(QLabel("입력"))
        layout.addWidget(self.input, 1)
        layout.addWidget(QLabel("결과"))
        layout.addWidget(self.output, 1)
        root.addWidget(content, 1)
        root.addLayout(bottom_action_bar(convert, copy))

    def words(self, text: str) -> list[str]:
        if "_" in text:
            return [part.lower() for part in text.split("_") if part]
        return [part.lower() for part in re.findall(r"[A-Z]?[a-z0-9]+|[A-Z]+(?=[A-Z]|$)", text)]

    def cycle_token(self, token: str) -> str:
        parts = self.words(token)
        if not parts:
            return token
        if token.upper() == token:
            return parts[0].lower() + "".join(part.capitalize() for part in parts[1:])
        if token[:1].islower() and "_" not in token:
            return "".join(part.capitalize() for part in parts)
        return "_".join(part.upper() for part in parts)

    def convert(self) -> None:
        text = self.input.toPlainText()
        result = re.sub(r"[A-Za-z][A-Za-z0-9_]*", lambda match: self.cycle_token(match.group(0)), text)
        self.output.setPlainText(result)
        self.input.setPlainText(result)


class JsonFormatterTab(QWidget):
    def __init__(self, main) -> None:
        super().__init__()
        self.main = main
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        content = QWidget()
        layout = QVBoxLayout(content)
        self.input = PlainTextEdit()
        self.output = PlainTextEdit()
        self.table_output = PlainTextEdit()
        convert = QPushButton("JSON 변환")
        excel = QPushButton("엑셀 양식 변환")
        copy = QPushButton("결과 복사")
        copy_table = QPushButton("엑셀 양식 복사")
        convert.clicked.connect(self.format_json)
        excel.clicked.connect(self.convert_to_excel_table)
        copy.clicked.connect(lambda: self.main.app.clipboard().setText(self.output.toPlainText()))
        copy_table.clicked.connect(lambda: self.main.app.clipboard().setText(self.table_output.toPlainText()))
        layout.addWidget(QLabel("JSON 원문"))
        layout.addWidget(self.input, 1)
        layout.addWidget(QLabel("가독화 결과"))
        layout.addWidget(self.output, 1)
        layout.addWidget(QLabel("엑셀 붙여넣기 양식"))
        layout.addWidget(self.table_output, 1)
        root.addWidget(content, 1)
        root.addLayout(bottom_action_bar(convert, excel, copy, copy_table))

    def parse_json(self):
        try:
            return json.loads(self.input.toPlainText())
        except Exception as exc:
            QMessageBox.warning(self, "JSON 오류", str(exc))
            return None

    def format_json(self) -> None:
        data = self.parse_json()
        if data is None:
            return
        self.output.setPlainText(json.dumps(data, ensure_ascii=False, indent=2))
        self.table_output.setPlainText(self.excel_table_text(data))

    def json_type(self, value) -> str:
        if value is None:
            return "null"
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return "number"
        if isinstance(value, str):
            return "string"
        if isinstance(value, list):
            return "array"
        if isinstance(value, dict):
            return "object"
        return type(value).__name__

    def split_path(self, path: str) -> tuple[str, str]:
        cleaned = re.sub(r"\[\d+\]", "[]", str(path or ""))
        if "." not in cleaned:
            return "", cleaned
        return cleaned.rsplit(".", 1)

    def display_value(self, value) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return str(value)
        if isinstance(value, str):
            return value
        if isinstance(value, list):
            return f"[array] {len(value)}"
        if isinstance(value, dict):
            return "{object}"
        return str(value)

    def flatten_rows(self, value, path: str = "") -> list[tuple[str, object]]:
        rows = []
        if path:
            rows.append((path, value))
        if isinstance(value, dict):
            for key, child in value.items():
                child_path = f"{path}.{key}" if path else str(key)
                rows.extend(self.flatten_rows(child, child_path))
        elif isinstance(value, list):
            for index, child in enumerate(value):
                child_path = f"{path}[{index}]" if path else f"[{index}]"
                rows.extend(self.flatten_rows(child, child_path))
        return rows

    def excel_table_text(self, data) -> str:
        lines = ["상위 경로\t필드명\t전체 경로\tTYPE\t값"]
        for path, value in self.flatten_rows(data):
            parent, field_name = self.split_path(path)
            lines.append("\t".join([
                parent,
                field_name,
                path,
                self.json_type(value),
                self.display_value(value).replace("\t", " ").replace("\r", " ").replace("\n", " "),
            ]))
        return "\n".join(lines)

    def convert_to_excel_table(self) -> None:
        data = self.parse_json()
        if data is None:
            return
        self.table_output.setPlainText(self.excel_table_text(data))


class TaggingReviewTab(QWidget):
    CHANNELS = ["WEB_PC", "WEB_MO", "APP_AOS", "APP_IOS"]
    TYPE_ALIASES = {
        "STRING": "string",
        "STR": "string",
        "TEXT": "string",
        "VARCHAR": "string",
        "NUMBER": "number",
        "NUMERIC": "number",
        "INT": "number",
        "INTEGER": "number",
        "FLOAT": "number",
        "DOUBLE": "number",
        "BOOLEAN": "boolean",
        "BOOL": "boolean",
        "ARRAY": "array",
        "LIST": "array",
        "OBJECT": "object",
        "JSON": "object",
    }

    def __init__(self, main) -> None:
        super().__init__()
        self.main = main
        self.definitions: list[dict] = []
        self.last_results: dict[str, list[dict]] = {}
        self.last_raw_inputs: dict[str, str] = {}
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        content = QWidget()
        layout = QVBoxLayout(content)
        self.definition_label = QLabel("정의서 미로드")
        self.definition_label.setObjectName("mutedText")
        self.definition_sample_label = QLabel("샘플 헤더: 필드명(한글) | 필드명(영어) | TYPE | SAMPLE")
        self.definition_sample_label.setObjectName("mutedText")
        load_btn = QPushButton("엑셀 정의서 불러오기")
        load_btn.clicked.connect(self.load_definition)
        top = QHBoxLayout()
        top.addWidget(load_btn)
        top.addWidget(self.definition_label, 1)
        top.addWidget(self.definition_sample_label, 2)
        layout.addLayout(top)
        grid = QGridLayout()
        self.inputs: dict[str, PlainTextEdit] = {}
        for index, channel in enumerate(self.CHANNELS):
            editor = PlainTextEdit()
            self.inputs[channel] = editor
            row = index // 2
            col = index % 2
            grid.addWidget(QLabel(channel), row * 2, col)
            grid.addWidget(editor, row * 2 + 1, col)
        layout.addLayout(grid, 3)
        self.output = PlainTextEdit()
        self.output.setReadOnly(True)
        layout.addWidget(QLabel("검토 요약"))
        layout.addWidget(self.output, 1)
        review_btn = QPushButton("검토")
        export_btn = QPushButton("결과 엑셀 추출")
        review_btn.clicked.connect(self.review)
        export_btn.clicked.connect(self.export_results)
        root.addWidget(content, 1)
        root.addLayout(bottom_action_bar(review_btn, export_btn))

    def normalize_header(self, value: str) -> str:
        return re.sub(r"\s+", "", str(value or "")).lower()

    def load_definition(self) -> None:
        if load_workbook is None:
            QMessageBox.warning(self, "의존성 필요", "openpyxl이 설치되어 있지 않습니다. requirements.txt 설치 후 다시 실행해 주세요.")
            return
        path, _ = QFileDialog.getOpenFileName(self, "태깅 정의서 불러오기", "", "Excel (*.xlsx *.xlsm)")
        if not path:
            return
        try:
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            headers = [self.normalize_header(cell.value) for cell in ws[1]]
            mapping = {}
            candidates = {
                "ko": ["필드명(한글)", "필드명한글", "한글필드명", "korean"],
                "en": ["필드명(영어)", "필드명영어", "영문필드명", "field", "key", "name", "eventproperty"],
                "type": ["type", "타입", "유형"],
                "sample": ["sample", "샘플", "예시"],
            }
            for key, names in candidates.items():
                normalized_names = {self.normalize_header(name) for name in names}
                for index, header in enumerate(headers):
                    if header in normalized_names:
                        mapping[key] = index
                        break
            if "en" not in mapping:
                raise ValueError("필드명(영어) 열을 찾을 수 없습니다.")
            definitions = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                field = str(row[mapping["en"]] or "").strip()
                if not field:
                    continue
                definitions.append({
                    "ko": str(row[mapping["ko"]] or "").strip() if "ko" in mapping else "",
                    "field": field,
                    "type": str(row[mapping["type"]] or "").strip() if "type" in mapping else "",
                    "sample": row[mapping["sample"]] if "sample" in mapping else "",
                })
            self.definitions = definitions
            self.definition_label.setText(f"{len(definitions)}개 필드 로드: {path}")
        except Exception as exc:
            QMessageBox.warning(self, "정의서 오류", str(exc))

    def normalize_type(self, value: str) -> str:
        raw = str(value or "").strip().upper()
        return self.TYPE_ALIASES.get(raw, raw.lower() if raw else "")

    def normalize_compare_name(self, value: str) -> str:
        return re.sub(r"[^0-9A-Za-z가-힣_]", "", str(value or "")).lower()

    def normalize_compare_path(self, value: str) -> str:
        cleaned = re.sub(r"\[\d+\]", "", str(value or ""))
        parts = [self.normalize_compare_name(part) for part in cleaned.split(".")]
        return ".".join(part for part in parts if part)

    def normalized_field_index(self, flattened: dict[str, object]) -> dict[str, list[str]]:
        index: dict[str, list[str]] = {}
        for field in flattened:
            normalized = self.normalize_compare_path(field)
            if normalized:
                index.setdefault(normalized, []).append(field)
            basename = self.normalize_compare_name(self.field_basename(field))
            if basename:
                index.setdefault(basename, []).append(field)
        return index

    def json_type(self, value) -> str:
        if value is None:
            return "null"
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return "number"
        if isinstance(value, str):
            return "string"
        if isinstance(value, list):
            return "array"
        if isinstance(value, dict):
            return "object"
        return type(value).__name__

    def flatten_json(self, value, prefix: str = "") -> dict[str, object]:
        result = {}
        if isinstance(value, dict):
            for key, child in value.items():
                path = f"{prefix}.{key}" if prefix else str(key)
                result[path] = child
                result.update(self.flatten_json(child, path))
        elif isinstance(value, list):
            for index, child in enumerate(value):
                path = f"{prefix}[{index}]"
                result[path] = child
                result.update(self.flatten_json(child, path))
        return result

    def empty_issue(self, value) -> str:
        if value is None:
            return "NULL"
        if value == "":
            return "빈 문자열"
        if isinstance(value, str) and value.strip() == "":
            return "공백 문자열"
        if isinstance(value, list) and not value:
            return "빈 배열"
        if isinstance(value, dict) and not value:
            return "빈 객체"
        return ""

    def compatible(self, expected: str, actual: str) -> bool:
        if not expected or actual == "null":
            return True
        return expected == actual

    def field_basename(self, field: str) -> str:
        return re.sub(r"\[\d+\]", "", str(field).split(".")[-1])

    def split_field_path(self, field: str) -> tuple[str, str]:
        cleaned = re.sub(r"\[\d+\]", "[]", str(field or ""))
        if "." not in cleaned:
            return "", cleaned
        return cleaned.rsplit(".", 1)

    def cell_value_text(self, value) -> str:
        if value == "":
            return ""
        if value is None:
            return "null"
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return str(value)
        if isinstance(value, str):
            return value
        return json.dumps(value, ensure_ascii=False)[:300]

    def row(
        self,
        status: str,
        field: str,
        expected_type: str = "",
        actual_type: str = "",
        value="",
        message: str = "",
        definition_field: str = "",
        definition_ko: str = "",
        sample="",
    ) -> dict:
        parent, field_name = self.split_field_path(field)
        return {
            "status": status,
            "message": message,
            "parent": parent,
            "field_name": field_name,
            "field": field,
            "definition_field": definition_field or field_name,
            "definition_ko": definition_ko,
            "expected_type": expected_type,
            "sample": self.cell_value_text(sample),
            "actual_type": actual_type,
            "value": self.cell_value_text(value),
        }

    def find_value(self, flattened: dict[str, object], field: str):
        if field in flattened:
            return field, flattened[field]
        index = self.normalized_field_index(flattened)
        normalized_path = self.normalize_compare_path(field)
        matches = list(dict.fromkeys(index.get(normalized_path, [])))
        if not matches:
            normalized_name = self.normalize_compare_name(field)
            matches = list(dict.fromkeys(index.get(normalized_name, [])))
        if len(matches) == 1:
            key = matches[0]
            return key, flattened[key]
        return "", None

    def warning_for_soft_type(self, expected: str, value) -> str:
        if not isinstance(value, str):
            return ""
        text = value.strip().lower()
        if expected == "number":
            try:
                float(text.replace(",", ""))
                return "숫자 TYPE인데 문자열 숫자로 수집되었습니다."
            except ValueError:
                return ""
        if expected == "boolean" and text in {"true", "false", "yes", "no", "y", "n", "0", "1"}:
            return "Boolean TYPE인데 문자열로 수집되었습니다."
        return ""

    def sample_type_issue(self, definition: dict) -> str:
        expected = self.normalize_type(definition.get("type", ""))
        sample = definition.get("sample")
        if sample in (None, "") or not expected:
            return ""
        actual = self.json_type(sample)
        if expected == "number" and isinstance(sample, str):
            try:
                float(sample.replace(",", ""))
                return ""
            except ValueError:
                pass
        return "" if self.compatible(expected, actual) else f"SAMPLE 타입이 TYPE과 다릅니다. 예상 {expected}, 샘플 {actual}"

    def array_type_issues(self, flattened: dict[str, object]) -> list[dict]:
        rows = []
        for field, value in flattened.items():
            if not isinstance(value, list) or len(value) < 2:
                continue
            types = Counter(self.json_type(item) for item in value)
            if len(types) > 1:
                rows.append(self.row("WARN", field, actual_type="array", value=value, message=f"배열 안에 여러 타입이 섞여 있습니다: {', '.join(sorted(types))}"))
        return rows

    def duplicate_json_key_issues(self, text: str) -> list[str]:
        duplicates = []

        def hook(pairs):
            counts = Counter(key for key, _value in pairs)
            duplicates.extend([key for key, count in counts.items() if count > 1])
            return dict(pairs)

        try:
            json.loads(text, object_pairs_hook=hook)
        except Exception:
            return []
        return sorted(set(duplicates))

    def parse_channel_json(self, channel: str):
        text = self.inputs[channel].toPlainText().strip()
        if not text:
            return None, "입력 없음"
        try:
            return json.loads(text), ""
        except Exception as exc:
            return None, str(exc)

    def review(self) -> None:
        if not self.definitions:
            QMessageBox.warning(self, "정의서 필요", "먼저 엑셀 태깅 정의서를 불러와 주세요.")
            return
        definitions_by_field: dict[str, dict] = {}
        definition_key_labels: dict[str, list[str]] = {}
        for item in self.definitions:
            key = self.normalize_compare_path(item["field"])
            if not key:
                continue
            definition_key_labels.setdefault(key, []).append(item["field"])
            definitions_by_field.setdefault(key, item)
        definition_fields = set(definitions_by_field)
        duplicate_fields = sorted(
            labels[0]
            for labels in definition_key_labels.values()
            if len(labels) > 1
        )
        known_types = set(self.TYPE_ALIASES.values())
        definition_warnings = []
        for definition in self.definitions:
            normalized_type = self.normalize_type(definition.get("type", ""))
            if normalized_type and normalized_type not in known_types:
                definition_warnings.append(self.row(
                    "WARN",
                    definition["field"],
                    expected_type=definition.get("type", ""),
                    message="정의서 TYPE이 표준 타입(String/Number/Boolean/Array/Object)이 아닙니다.",
                    definition_field=definition.get("field", ""),
                    definition_ko=definition.get("ko", ""),
                    sample=definition.get("sample", ""),
                ))
            sample_issue = self.sample_type_issue(definition)
            if sample_issue:
                definition_warnings.append(self.row(
                    "WARN",
                    definition["field"],
                    expected_type=definition.get("type", ""),
                    value=definition.get("sample"),
                    message=sample_issue,
                    definition_field=definition.get("field", ""),
                    definition_ko=definition.get("ko", ""),
                    sample=definition.get("sample", ""),
                ))
        results: dict[str, list[dict]] = {}
        raw_inputs: dict[str, str] = {}
        summary_lines = []
        for channel in self.CHANNELS:
            raw_text = self.inputs[channel].toPlainText().strip()
            raw_inputs[channel] = raw_text
            data, error = self.parse_channel_json(channel)
            rows = list(definition_warnings)
            if error:
                rows.append(self.row("ERROR", "", message=error))
                results[channel] = rows
                summary_lines.append(f"{channel}: JSON 오류")
                continue
            for key in self.duplicate_json_key_issues(raw_text):
                rows.append(self.row("WARN", key, message="JSON 객체 안에 중복 키가 있습니다. 마지막 값만 검토에 사용됩니다."))
            flattened = self.flatten_json(data)
            actual_fields = set(flattened)
            for _definition_key, definition in definitions_by_field.items():
                field = definition["field"]
                actual_field, value = self.find_value(flattened, field)
                if not actual_field:
                    rows.append(self.row(
                        "MISSING",
                        field,
                        expected_type=definition.get("type", ""),
                        message="정의서 필드가 JSON에 없습니다.",
                        definition_field=field,
                        definition_ko=definition.get("ko", ""),
                        sample=definition.get("sample", ""),
                    ))
                    continue
                expected = self.normalize_type(definition.get("type", ""))
                actual = self.json_type(value)
                empty = self.empty_issue(value)
                if not self.compatible(expected, actual):
                    soft_message = self.warning_for_soft_type(expected, value)
                    status = "WARN" if soft_message else "TYPE_MISMATCH"
                    message = soft_message or f"예상 {expected}, 실제 {actual}"
                elif empty:
                    status = "EMPTY"
                    message = empty
                else:
                    status = "OK"
                    message = ""
                rows.append(self.row(
                    status,
                    actual_field,
                    expected_type=definition.get("type", ""),
                    actual_type=actual,
                    value=value,
                    message=message,
                    definition_field=field,
                    definition_ko=definition.get("ko", ""),
                    sample=definition.get("sample", ""),
                ))
            for field in sorted(actual_fields):
                if re.search(r"\[\d+\]$", field):
                    continue
                normalized_field = self.normalize_compare_path(field)
                normalized_basename = self.normalize_compare_name(self.field_basename(field))
                if normalized_field in definition_fields or normalized_basename in definition_fields:
                    continue
                value = flattened[field]
                if isinstance(value, (dict, list)):
                    continue
                rows.append(self.row("UNDEFINED", field, actual_type=self.json_type(value), value=value, message="정의서에 없는 JSON 키입니다."))
            rows.extend(self.array_type_issues(flattened))
            if duplicate_fields:
                rows.append(self.row("WARN", ", ".join(duplicate_fields), message="정의서에 중복 필드가 있습니다."))
            results[channel] = rows
            bad_count = sum(1 for row in rows if row["status"] != "OK")
            summary_lines.append(f"{channel}: OK {len(rows) - bad_count} / 확인필요 {bad_count}")
        self.last_results = results
        self.last_raw_inputs = raw_inputs
        self.output.setPlainText("\n".join(summary_lines))

    def result_headers(self) -> list[str]:
        return [
            "상태",
            "메시지",
            "필드명(한글)",
            "필드명(영어)",
            "정의 TYPE",
            "SAMPLE",
            "상위 경로",
            "필드명",
            "JSON TYPE",
            "JSON 값",
        ]

    def export_message(self, row: dict) -> str:
        status = str(row.get("status", "") or "").upper()
        message = str(row.get("message", "") or "")
        if status == "OK":
            return ""
        if status == "MISSING":
            return "JSON 없음"
        if status == "UNDEFINED":
            return "정의서 없음"
        if status == "TYPE_MISMATCH":
            expected = self.normalize_type(row.get("expected_type", ""))
            actual = str(row.get("actual_type", "") or "")
            return f"타입 불일치: {expected}->{actual}" if expected or actual else "타입 불일치"
        if status == "EMPTY":
            return "빈 값"
        if status == "ERROR":
            return "JSON 오류"
        if status == "WARN":
            if "중복 키" in message:
                return "JSON 키 중복"
            if "중복 필드" in message:
                return "정의서 중복"
            if "표준 타입" in message:
                return "TYPE 확인"
            if "SAMPLE" in message:
                return "SAMPLE 타입 불일치"
            if "여러 타입" in message:
                return "배열 타입 혼재"
            if "문자열 숫자" in message:
                return "문자열 숫자"
            if "문자열로 수집" in message:
                return "문자열 Boolean"
        return message

    def result_values(self, row: dict) -> list:
        return [
            row.get("status", ""),
            self.export_message(row),
            row.get("definition_ko", ""),
            row.get("definition_field", ""),
            row.get("expected_type", ""),
            row.get("sample", ""),
            row.get("parent", ""),
            row.get("field_name", ""),
            row.get("actual_type", ""),
            row.get("value", ""),
        ]

    def status_fill(self, status: str):
        if PatternFill is None:
            return None
        colors = {
            "MISSING": "DDEBFF",
            "UNDEFINED": "E2F7E1",
            "OK": None,
        }
        color = colors.get(str(status or "").upper(), "FFE1E1")
        return PatternFill("solid", fgColor=color) if color else None

    def status_font(self, status: str):
        if Font is None:
            return None
        colors = {
            "MISSING": "FF2563EB",
            "UNDEFINED": "FF15803D",
            "OK": None,
        }
        color = colors.get(str(status or "").upper(), "FFDC2626")
        return Font(color=color, bold=True, sz=9) if color else None

    def apply_status_fill(self, ws, row_index: int, status: str, start_col: int = 1, end_col: int | None = None) -> None:
        fill = self.status_fill(status)
        if fill is None:
            return
        end_col = end_col or ws.max_column
        for col in range(start_col, end_col + 1):
            ws.cell(row=row_index, column=col).fill = fill

    def apply_status_font(self, ws, row_index: int, status: str, start_col: int = 1, end_col: int | None = None) -> None:
        font = self.status_font(status)
        if font is None:
            return
        end_col = end_col or ws.max_column
        for col in range(start_col, end_col + 1):
            ws.cell(row=row_index, column=col).font = font

    def style_header(self, ws, row_index: int = 1) -> None:
        if Font is None or PatternFill is None:
            return
        header_fill = PatternFill("solid", fgColor="F3F4F6")
        for cell in ws[row_index]:
            cell.font = Font(bold=True, sz=9)
            cell.fill = header_fill

    def write_grouped_result_header(self, ws, include_channel: bool = False) -> int:
        headers = self.result_headers()
        col = 1
        if include_channel:
            ws.cell(row=1, column=col).value = "채널명"
            col += 1
        for index, header in enumerate(headers, start=col):
            ws.cell(row=1, column=index).value = header
        self.style_header(ws, 1)
        if PatternFill is not None:
            group_fills = {
                "status": PatternFill("solid", fgColor="FCE7F3"),
                "definition": PatternFill("solid", fgColor="DBEAFE"),
                "json": PatternFill("solid", fgColor="DCFCE7"),
                "channel": PatternFill("solid", fgColor="F3F4F6"),
            }
            if include_channel:
                ws.cell(row=1, column=1).fill = group_fills["channel"]
                base = 2
            else:
                base = 1
            for idx in range(base, base + 2):
                ws.cell(row=1, column=idx).fill = group_fills["status"]
            for idx in range(base + 2, base + 6):
                ws.cell(row=1, column=idx).fill = group_fills["definition"]
            for idx in range(base + 6, base + 10):
                ws.cell(row=1, column=idx).fill = group_fills["json"]
        if Alignment is not None:
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return 2

    def autosize_sheet(self, ws, max_width: int = 42) -> None:
        if get_column_letter is None:
            return
        for col_index, column in enumerate(ws.columns, start=1):
            letter = get_column_letter(col_index)
            width = 8
            for cell in column:
                width = max(width, min(max_width, len(str(cell.value or "").split("\n")[0]) + 2))
            ws.column_dimensions[letter].width = width

    def apply_workbook_font(self, wb) -> None:
        if Font is None:
            return
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    font = copy(cell.font)
                    font.sz = 9
                    cell.font = font

    def pretty_raw_json(self, raw_text: str) -> str:
        if not raw_text:
            return ""
        try:
            return json.dumps(json.loads(raw_text), ensure_ascii=False, indent=2)
        except Exception:
            return raw_text

    def channel_has_export_input(self, channel: str) -> bool:
        raw = str(self.last_raw_inputs.get(channel, "") or "").strip()
        return bool(raw) and raw.lower() not in {"null", "none"}

    def issue_matches_raw_line(self, issue: dict, line: str) -> bool:
        tokens = [
            str(issue.get("field_name") or ""),
            str(issue.get("definition_field") or ""),
            str(issue.get("field") or "").split(".")[-1],
        ]
        return any(token and (f'"{token}"' in line or token in line) for token in tokens)

    def write_raw_sheet(self, wb, channel: str, rows: list[dict]) -> None:
        ws = wb.create_sheet(f"RAW_{channel}")
        ws.cell(row=1, column=1).value = "원문 JSON"
        raw_lines = self.pretty_raw_json(self.last_raw_inputs.get(channel, "")).splitlines() or [""]
        issue_rows = [row for row in rows if row.get("status") != "OK"]
        self.style_header(ws, 1)
        for row_index, line in enumerate(raw_lines, start=2):
            cell = ws.cell(row=row_index, column=1)
            cell.value = line
            matched = next((issue for issue in issue_rows if self.issue_matches_raw_line(issue, line)), None)
            if matched:
                self.apply_status_font(ws, row_index, matched.get("status", ""), 1, 1)
        self.autosize_sheet(ws)
        ws.column_dimensions["A"].width = 100

    def export_results(self) -> None:
        if not self.last_results:
            self.review()
            if not self.last_results:
                return
        if Workbook is None:
            QMessageBox.warning(self, "의존성 필요", "openpyxl이 설치되어 있지 않습니다. requirements.txt 설치 후 다시 실행해 주세요.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "검토 결과 저장", "tagging_review_result.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        wb = Workbook()
        summary = wb.active
        summary.title = "Summary"
        summary_start_row = self.write_grouped_result_header(summary, include_channel=True)
        export_items = [
            (channel, rows)
            for channel, rows in self.last_results.items()
            if self.channel_has_export_input(channel)
        ]
        if not export_items:
            QMessageBox.warning(self, "입력 없음", "엑셀로 추출할 JSON 입력 채널이 없습니다.")
            return
        for channel, rows in export_items:
            for row in rows:
                if row.get("status") == "OK":
                    continue
                summary.append([channel] + self.result_values(row))
                self.apply_status_fill(summary, summary.max_row, row.get("status", ""), 1, 1 + len(self.result_headers()))
            ws = wb.create_sheet(channel)
            detail_start_row = self.write_grouped_result_header(ws, include_channel=False)
            for row in rows:
                ws.append(self.result_values(row))
                self.apply_status_fill(ws, ws.max_row, row.get("status", ""), 1, len(self.result_headers()))
            ws.freeze_panes = f"A{detail_start_row}"
            self.autosize_sheet(ws)
            self.write_raw_sheet(wb, channel, rows)
        summary.freeze_panes = f"A{summary_start_row}"
        self.autosize_sheet(summary)
        self.apply_workbook_font(wb)
        wb.save(path)
        QMessageBox.information(self, "저장 완료", "검토 결과 엑셀을 저장했습니다.")


class TextToolsTab(QWidget):
    def __init__(self, main) -> None:
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        tabs = QTabWidget()
        tabs.addTab(UrlCodecTab(main), "URL 인코딩")
        tabs.addTab(UtmTab(main), "UTM")
        tabs.addTab(LineBreakTab(main), "줄바꿈/따옴표")
        tabs.addTab(CaseCycleTab(main), "대소문자 변환")
        tabs.addTab(TaggingReviewTab(main), "태깅 검토")
        tabs.addTab(JsonFormatterTab(main), "JSON 변환기")
        layout.addWidget(tabs, 1)
